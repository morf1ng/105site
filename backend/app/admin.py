from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Role, User, Course, CourseRegistration
from .schemas import Token
from .utils import hash_password
from .database import get_session
from .dependencies import role_required

router = APIRouter(prefix="/admin", tags=["Admin"])



@router.get("/roles")
def get_roles(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    return session.query(Role).all()

@router.post("/roles")
def create_role(name: str = Form(...), session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    if session.query(Role).filter(Role.name == name).first():
        raise HTTPException(status_code=400, detail="Role already exists")
    role = Role(name=name)
    session.add(role)
    session.commit()
    session.refresh(role)
    return role

@router.put("/roles/{role_id}")
def update_role(role_id: int, name: str = Form(...), session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    role = session.query(Role).get(role_id)
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    role.name = name
    session.commit()
    return role

@router.delete("/roles/{role_id}")
def delete_role(role_id: int, session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    role = session.query(Role).get(role_id)
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Проверяем, используется ли эта роль
    users_with_role = session.query(User).filter(User.role_ids.like(f"%{role_id}%")).all()
    if users_with_role:
        raise HTTPException(status_code=403, detail="Cannot delete a role that still has users")
    
    session.delete(role)
    session.commit()
    return {"detail": "Deleted"}



@router.get("/users")
def get_users(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    return session.query(User).all()

@router.get("/users/{user_id}")
def get_user_detail(user_id: int, session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    user_obj = session.query(User).get(user_id)
    if not user_obj:
        raise HTTPException(status_code=404, detail="User not found")
    return user_obj

@router.post("/users")
def create_user(
    email: str = Form(...),
    password: str = Form(...),
    fullname: str = Form(None),
    role_ids: str = Form(...),  # Передаем как строку с разделителями, например "1,2,3"
    session: Session = Depends(get_session),
    user=Depends(role_required("admin")),
):
    if session.query(User).filter(User.email == email).first():
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Парсим role_ids
    try:
        role_id_list = [int(rid.strip()) for rid in role_ids.split(",") if rid.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid role_ids format")
    
    if not role_id_list:
        raise HTTPException(status_code=400, detail="At least one role is required")
    
    # Проверяем существование всех ролей
    db_roles = session.query(Role).filter(Role.id.in_(role_id_list)).all()
    if len(db_roles) != len(role_id_list):
        raise HTTPException(status_code=400, detail="One or more roles do not exist")
    
    password_hash = hash_password(password)
    new_user = User(email=email, fullname=fullname, password_hash=password_hash, role_ids=role_ids)
    session.add(new_user)
    session.commit()
    session.refresh(new_user)
    return new_user

@router.put("/users/{user_id}")
def update_user(
    user_id: int,
    email: str = Form(None),
    password: str = Form(None),
    fullname: str = Form(None),
    role_ids: str = Form(None),  # Передаем как строку с разделителями, например "1,2,3"
    session: Session = Depends(get_session),
    user=Depends(role_required("admin")),
):
    user_obj = session.query(User).get(user_id)
    if not user_obj:
        raise HTTPException(status_code=404, detail="User not found")

    # нельзя убрать роль admin у последнего админа
    admin_role = session.query(Role).filter(Role.name == "admin").first()
    if admin_role:
        current_role_ids = user_obj.get_role_ids_list()
        has_admin_role = admin_role.id in current_role_ids
        
        # Считаем всех админов
        all_users = session.query(User).all()
        admin_count = sum(1 for u in all_users if admin_role.id in u.get_role_ids_list())
        
        if has_admin_role and admin_count == 1 and role_ids is not None:
            # Парсим новые роли
            try:
                role_id_list = [int(rid.strip()) for rid in role_ids.split(",") if rid.strip()]
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid role_ids format")
            
            if admin_role.id not in role_id_list:
                raise HTTPException(403, detail="Cannot remove admin role from the last admin")

    if email is not None:
        user_obj.email = email
    if fullname is not None:
        user_obj.fullname = fullname
    if password is not None:
        user_obj.password_hash = hash_password(password)
    if role_ids is not None:
        try:
            role_id_list = [int(rid.strip()) for rid in role_ids.split(",") if rid.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid role_ids format")
        
        if not role_id_list:
            raise HTTPException(status_code=400, detail="At least one role is required")
        
        db_roles = session.query(Role).filter(Role.id.in_(role_id_list)).all()
        if len(db_roles) != len(role_id_list):
            raise HTTPException(status_code=400, detail="One or more roles do not exist")
        
        user_obj.role_ids = role_ids

    session.commit()
    return user_obj

@router.delete("/users/{user_id}")
def delete_user(user_id: int, session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    user_obj = session.query(User).get(user_id)
    if not user_obj:
        raise HTTPException(status_code=404, detail="User not found")

    # нельзя удалить последнего админа
    admin_role = session.query(Role).filter(Role.name == "admin").first()
    if admin_role:
        current_role_ids = user_obj.get_role_ids_list()
        has_admin_role = admin_role.id in current_role_ids
        
        # Считаем всех админов
        all_users = session.query(User).all()
        admin_count = sum(1 for u in all_users if admin_role.id in u.get_role_ids_list())
        
        if has_admin_role and admin_count == 1:
            raise HTTPException(403, detail="Cannot delete last admin")

    session.delete(user_obj)
    session.commit()
    return {"detail": "Deleted"}


# --- Курсы (админка) ---
@router.get("/courses")
def get_courses_admin(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    return session.query(Course).order_by(Course.id).all()

@router.post("/courses")
def create_course_admin(title: str = Form(...), session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    course = Course(title=title)
    session.add(course)
    session.commit()
    session.refresh(course)
    return course

@router.put("/courses/{course_id}")
def update_course_admin(course_id: int, title: str = Form(...), session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    course = session.query(Course).get(course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Курс не найден")
    course.title = title
    session.commit()
    session.refresh(course)
    return course

@router.delete("/courses/{course_id}")
def delete_course_admin(course_id: int, session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    course = session.query(Course).get(course_id)
    if not course:
        raise HTTPException(status_code=404, detail="Курс не найден")
    session.delete(course)
    session.commit()
    return {"detail": "Deleted"}

@router.get("/course-registrations")
def get_course_registrations(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Список всех заявок на курсы."""
    regs = session.query(CourseRegistration).order_by(CourseRegistration.created_at.desc()).all()
    return [{"id": r.id, "full_name": r.full_name, "phone": r.phone, "email": r.email, "course_id": r.course_id, "course_title": r.course.title if r.course else None, "support_type": r.support_type, "created_at": str(r.created_at) if r.created_at else None} for r in regs]


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "ФИО", "Роли", "Создан"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email)
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=u.role_ids or "")
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/course-registrations", response_class=StreamingResponse)
def export_course_registrations_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт заявок на курсы в Excel."""
    regs = session.query(CourseRegistration).order_by(CourseRegistration.created_at.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки на курсы"
    headers = ["ID", "ФИО", "Телефон", "Email", "Курс", "Тип", "Дата регистрации"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, r in enumerate(regs, 2):
        ws.cell(row=row, column=1, value=r.id)
        ws.cell(row=row, column=2, value=r.full_name)
        ws.cell(row=row, column=3, value=r.phone)
        ws.cell(row=row, column=4, value=r.email)
        ws.cell(row=row, column=5, value=r.course.title if r.course else "")
        ws.cell(row=row, column=6, value="С поддержкой" if r.support_type == "with_support" else "Базовый")
        ws.cell(row=row, column=7, value=str(r.created_at) if r.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=course_registrations.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users_list = session.query(User).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "ФИО", "Роли", "Создан"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users_list, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email or "")
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=u.role_ids or "")
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).order_by(User.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "Имя", "Роли", "Создан"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email)
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=u.role_ids or "")
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).order_by(User.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "Имя", "Роли", "Создан"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email or "")
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=u.role_ids or "")
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 25
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).order_by(User.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "Имя", "ID ролей", "Дата создания"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email or "")
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=u.role_ids or "")
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).order_by(User.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "Имя", "Роли", "Создан", "Обновлён"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email)
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=u.role_ids or "")
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
        ws.cell(row=row, column=6, value=str(u.updated_at) if u.updated_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).order_by(User.id).all()
    roles = {r.id: r.name for r in session.query(Role).all()}
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "Имя", "Роли", "Дата создания"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        role_names = ", ".join(roles.get(rid, str(rid)) for rid in u.get_role_ids_list())
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email)
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=role_names)
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.get("/export/users", response_class=StreamingResponse)
def export_users_xlsx(session: Session = Depends(get_session), user=Depends(role_required("admin"))):
    """Экспорт пользователей в Excel."""
    users = session.query(User).order_by(User.id).all()
    roles = {r.id: r.name for r in session.query(Role).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    headers = ["ID", "Email", "ФИО", "Роли", "Дата создания"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row, u in enumerate(users, 2):
        role_names = ", ".join(roles.get(rid, str(rid)) for rid in u.get_role_ids_list())
        ws.cell(row=row, column=1, value=u.id)
        ws.cell(row=row, column=2, value=u.email)
        ws.cell(row=row, column=3, value=u.fullname or "")
        ws.cell(row=row, column=4, value=role_names)
        ws.cell(row=row, column=5, value=str(u.created_at) if u.created_at else "")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )